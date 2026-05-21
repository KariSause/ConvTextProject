import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from drive import download as drive_download


RED_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

def is_valid_xlsx(path: Path) -> bool:
    try:
        if not path.exists():
            return False
        with open(path, "rb") as f:
            return f.read(4) == b"PK\x03\x04"
    except Exception:
        return False


def download_template(template_path: Path, file_id: str):
    template_path.parent.mkdir(parents=True, exist_ok=True)
    print(f"[Excel] Downloading from drive -> {template_path}")
    drive_download(file_id, template_path.name, str(template_path.parent))
    
    if not is_valid_xlsx(template_path):
        raise RuntimeError(
            "File is not valid"
        )

    if template_path.stat().st_size < 1000:
        raise RuntimeError("File is too small, likely corrupted")

def append_to_excel(template_path, output_path, single_analysis):
    template_path = Path(template_path)
    output_path = Path(output_path)
    
    file_to_open = output_path if output_path.exists() else template_path
    
    if not file_to_open.exists():
        raise FileNotFoundError(f"Не знайдено жодного Excel файлу: {file_to_open}")
        
    wb = load_workbook(str(file_to_open))
    ws = wb.active

    current_row = ws.max_row + 1 if ws.max_row > 2 else 3
    row = single_analysis
    
    ws.cell(row=current_row, column=1, value=row.get("date"))             # A (1) - Дата
    ws.cell(row=current_row, column=2, value=row.get("call_type"))        # B (2) - Тип звернення
    ws.cell(row=current_row, column=6, value=row.get("script"))           # F (6) - Початок розмови, представлення
    ws.cell(row=current_row, column=7, value=row.get("car_info"))         # G (7) - Дізнався кузов
    ws.cell(row=current_row, column=8, value=row.get("car_info"))         # H (8) - Дізнався рік
    ws.cell(row=current_row, column=9, value=row.get("car_info"))         # I (9) - Дізнався пробіг
    ws.cell(row=current_row, column=10, value=row.get("upsell"))          # J (10) - Пропозиція комплексної діагностики
    ws.cell(row=current_row, column=11, value=row.get("service_history")) # K (11) - Дізнався які роботи робились раніше
    ws.cell(row=current_row, column=13, value=row.get("closing"))         # M (13) - Завершення розмови, прощання

    ws.cell(row=current_row, column=17, value=row.get("call_result"))     # Q (17) - Результат
    ws.cell(row=current_row, column=19, value=row.get("parts"))           # S (19) - Запчастини
    ws.cell(row=current_row, column=20, value=row.get("comment"))         # T (20) - Коментар


    if row.get("bad_call"):
        ws.cell(row=current_row, column=20).fill = RED_FILL
    ws.cell(row=current_row, column=18, value=f"=SUM(F{current_row}:K{current_row})+M{current_row}") 

    # Зберігаємо та закриваємо файл
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()
    
    print(f"[Excel] Дані успішно додано в рядок {current_row}")